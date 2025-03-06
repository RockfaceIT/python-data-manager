import openpyxl

class MyWorksheet:

    ## define input/output files & activate sheet
    def __init__(self, input_file, output_file, column_headings, columns_delete):

        ## Set output file
        self.output_file = output_file

        ## Open and activate workbook
        self.wb = openpyxl.load_workbook(input_file)
        self.sheet = self.wb.active

        ## Set headings and delete columns
        self.column_headings = column_headings
        self.columns_delete = columns_delete

    def get_value(self, row, column):
        return self.sheet.cell(row=row, column=column).value

    def set_value(self, row, column, value):
        self.sheet.cell(row=row, column=column).value = value

    def set_fullname(self, first_name_col, last_name_col):

        for row in range(2, self.sheet.max_row + 1):
            ## Set Full Name from Firstname and Lastname fields
            self.set_value(row, first_name_col, self.get_value(row, first_name_col) + " " +
                           self.get_value(row, last_name_col))

    def delete_columns(self):

        ## Remove unused columns
        for column, length in self.columns_delete.items():
            self.sheet.delete_cols(column, length)

    def manage_columns(self):

        ## Remove unused columns
        self.delete_columns()

        ## Set Column Names for Top Row
        for col_num, heading in enumerate(self.column_headings, 1):
            cell = self.sheet.cell(row=1, column=col_num)
            cell.value = heading

    def save_file(self):
        self.wb.save(self.output_file)
