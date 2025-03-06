import openpyxl

## Load File
wb = openpyxl.load_workbook("input_file.xlsx")
save_filename = "output_file.xlsx"

## Get Active sheet
sheet = wb.active

## adjust data

## Loop over Rows and merge first and last names
for row in range(2, sheet.max_row + 1):

    ## Set Full Name from Firstname and Lastname fields
    sheet.cell(row=row, column=3).value = sheet.cell(row=row, column=3).value + " " + sheet.cell(row=row, column=4).value


## Remove unused columns
sheet.delete_cols(1,1)
sheet.delete_cols(3,3)
sheet.delete_cols(5,1)

## Set Column Names for JotForm
sheet["A1"] = "ID"
sheet["B1"] = "Provider's Name"
sheet["C1"] = "California Acupuncture License Number"
sheet["D1"] = "Personal Email"

## Save file
wb.save(save_filename)

